from django.shortcuts import render, redirect
from .form import CreateUserForm, filtreUserForm, filtreTrajectForm, filtreAllForm, UpdateUserForm, UpdateUserPasswordForm, CreateMissionForm, CreateMessageForm, uploadImageForm
from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from .serializer import LoginSerializer, RapportSerializer, UserSerializer_image, PositionSerailize, RapportListSerializer, RegisterSerializer, missionSerializer, messageSerializer, UserSerializer
from rest_framework.response import Response
from rest_framework import generics
from rest_framework.views import APIView
from .models import Rapport, User, HistoriqueConnexion, Position, mission, message
from rest_framework.generics import GenericAPIView
from rest_framework.permissions import AllowAny
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from django.http import HttpResponse
from django.utils import timezone
import json
from django.urls import reverse_lazy
from datetime import datetime
# Create your views here.

#page de la map
@login_required(login_url=reverse_lazy('connection'))
def map(request):
    
    user = Position.objects.all()
     # Convertir les valeurs de latitude et longitude en nombres décimaux
    user_data = [
        {
            "id": user.id,
            "latitude": float(user.latitude),
            "longitude": float(user.longitude),
            "name": user.utilisateur.username,
            "matricule" : user.utilisateur.matricule
        }
        for user in user
    ]
    context = {
        "user_data": json.dumps(user_data)  # Convertir la liste en chaîne JSON
    }

    return render(request, 'index.html', context)


#page d'insription 
def inscription_page(request):
    register_form = CreateUserForm()
    if request.method == 'POST' or 'FILES':
        register_form = CreateUserForm(data=request.POST ,files=request.FILES)
        if register_form.is_valid() :
            register_form.save()
            messages.info(request, "Account Created Successfully!")
            return redirect('inscription')
        else:
            messages.error(request, "mot de passe ou nom incorrect")
    else:
        print('mauvais')

    return render(request, 'inscription.html', {'register_form': register_form})

#pour filtrer les info de tous les user

def info_all_user(request):

    form = filtreAllForm(request.GET)
    user = form.filter() if form.is_valid() else User.objects.all()
    context = {
        'form' : form,
        'user' : user
    }
    return render(request, 'all_info_user.html', context)

#page de connection 
def connection_page(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user =  authenticate(request, username=username, password=password )
        if user is not None:
            login(request, user)
            return redirect('gestion')
        else:
            messages.info(request, "Identifiants invalide")   
    return render(request, 'connection.html', {})

#pour les rapport 
@login_required(login_url=reverse_lazy('connection'))
def rapport_views(request) :
    user = request.user
    rapport = Rapport.objects.filter(receveur = user)
    rapport_admin = Rapport.objects.all()
    context = {
        'rapport' : rapport,
        'rapport_admin' : rapport_admin
    }
    
    return render(request, 'rapport.html', context)
    

#pour les gestion 
@login_required(login_url=reverse_lazy('connection'))
def gestion(request) : 
    user = User.objects.all()
    connecte = request.user
    
    context = {
        "user" : user,
        "connecte" : connecte
    }

    return render(request, 'gestion.html', context)


@login_required(login_url=reverse_lazy('connection'))
def update_user(request, my_id):
    obj = get_object_or_404(User, id=my_id)
    form = UpdateUserForm(request.POST or None, instance=obj)
    message = ''


    if request.method == 'POST' or None:
        if form.is_valid():
            user_instance = form.save(commit=False)
            
            user_instance.save()
            return redirect('gestion')
        else:
            print(form.errors)
    else:
        print("pas post")

    return render(request, 'update_user.html', {'register_form': form, 'message': message}) 

@login_required(login_url=reverse_lazy('connection'))
def update_user_password(request, my_id):
    obj = get_object_or_404(User, id=my_id)
    form = UpdateUserPasswordForm(request.POST or None, instance=obj)
    message = ''

    if request.method == 'POST' or None:
        if form.is_valid():
            user_instance = form.save(commit=False)

            # Vérifiez si un nouveau mot de passe est fourni dans le formulaire
            new_password = form.cleaned_data.get('password')
            if new_password:
                # Utilisez set_password pour hasher le nouveau mot de passe
                user_instance.set_password(new_password)
            # ... (votre code pour mettre à jour les autres champs)
            user_instance.save()
            return redirect('gestion')
        else:
            print(form.errors)
    else:
        print("pas post")

    return render(request, 'update_password.html', {'register_form': form, 'message': message}) 



@csrf_exempt
def delete_user(request, my_id):
    obj = get_object_or_404(User, id=my_id)
    name = obj.username
    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)  # Réponse sans contenu (No Content)

    return render(request, 'delete_user.html', {"name": name})

@login_required(login_url=reverse_lazy('connection'))
def time_connection(request):
    form = filtreUserForm(request.GET)
    persons = form.filter() if form.is_valid() else HistoriqueConnexion.objects.all()
    context = {
        'persons': persons,
        'form': form,
    }
    return render(request, 'connection_time.html', context)

def download_excel_connection(request):

    persons = HistoriqueConnexion.objects.all()

    wb = Workbook()

    ws = wb.active

    ws.append(['Noms', 'Heure de connexion'])

    for person in persons:
        # Utilisez timezone.make_aware pour rendre la date consciente du fuseau horaire
        aware_date = timezone.make_aware(person.date_connexion)
        # Utilisez timezone.localtime pour convertir la date dans le fuseau horaire local
        local_time = timezone.localtime(aware_date)
        # Formatez la date comme vous le souhaitez (par exemple, '5 Oct 2023, 18:14')
        formatted_date = local_time.strftime('%d-%b-%Y %H:%M:%S')

        ws.append([person.utilisateur.username, formatted_date])

    # Créez une réponse HTTP avec le type de contenu approprié pour un fichier Excel (.xlsx)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Définissez l'en-tête Content-Disposition pour indiquer le nom du fichier téléchargé
    response['Content-Disposition'] = 'attachment; filename=historique_connexion.xlsx'

    # Enregistrez le classeur dans la réponse HTTP
    wb.save(response)

    return response

def supprimer_historique_connexion(request):
    if request.method == 'POST':
        # Supprimez les données de la table HistoriqueConnexion
        HistoriqueConnexion.objects.all().delete()
        return JsonResponse({'message': 'Historique de connexion supprimé avec succès.'})

    return render(request, 'connection_time.html')

from django.utils import timezone

def download_rapport(request):
    rapports = Rapport.objects.all()

    wb = Workbook()

    ws = wb.active

    ws.append(['Envoyeur', 'Receveur', 'Texte', 'Note', 'Heure'])

    for rapport in rapports:
        # Utilisez timezone.make_aware pour rendre la date consciente du fuseau horaire
        aware_date = timezone.make_aware(rapport.heure)
        # Utilisez timezone.localtime pour convertir la date dans le fuseau horaire local
        local_time = timezone.localtime(aware_date)
        # Formatez la date comme vous le souhaitez (par exemple, '5 Oct 2023, 18:14')
        formatted_date = local_time.strftime('%d-%b-%Y %H:%M:%S')

        ws.append([rapport.envoyeur.username, rapport.receveur.username, rapport.texte, rapport.note, formatted_date])

    # Créez une réponse HTTP avec le type de contenu approprié pour un fichier Excel (.xlsx)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Définissez l'en-tête Content-Disposition pour indiquer le nom du fichier téléchargé
    response['Content-Disposition'] = 'attachment; filename=historique_rapport.xlsx'

    # Enregistrez le classeur dans la réponse HTTP
    wb.save(response)

    return response


def time_traject(request):
    form = filtreTrajectForm(request.GET)
    traject = form.filter() if form.is_valid() else Position.objects.all()
    context = {
        'traject': traject,
        'form': form,
    }
    return render(request, 'heure_parcours.html', context)    

def download_traject(request):

    traject = Position.objects.all()

    wb = Workbook()

    ws = wb.active

    ws.append(['personne','localite', 'heure'])

    for traject in traject:

        local_time = timezone.localtime(traject.date)

        formatted_date = local_time.strftime('%d-%b-%Y %H:%M:%S')

        ws.append([traject.utilisateur.username, traject.localite, formatted_date])

    # Créez une réponse HTTP avec le type de contenu approprié pour un fichier Excel (.xlsx)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Définissez l'en-tête Content-Disposition pour indiquer le nom du fichier téléchargé
    response['Content-Disposition'] = 'attachment; filename=historique_traject.xlsx'

    # Enregistrez le classeur dans la réponse HTTP
    wb.save(response)

    return response

def supprimer_traject(request):
    if request.method == 'POST':
        # Supprimez les données de la table HistoriqueConnexion
        Position.objects.all().delete()
        return JsonResponse({'message': 'Historique de connexion supprimé avec succès.'})

    return render(request, 'heure_parcours.html')

def supprimer_rapport(request):
    if request.method == 'POST':
        Rapport.objects.all().delete()
        return JsonResponse({'message': 'Historique de rapport supprimé avec succès.'})

    return render(request, 'rapport.html')

def logout_user(request):
    logout(request)
    return redirect('connection')

def create_mission(request):
    if request.method == 'POST':
        form = CreateMissionForm(request.POST)
        if form.is_valid():
            # Créer une instance de mission à partir des données du formulaire
            new_mission = form.save(commit=False)

            # Assurez-vous que les champs debut et fin sont correctement formatés avant de sauvegarder
            debut_str = request.POST.get('debut', '')
            fin_str = request.POST.get('fin', '')

            debut = datetime.strptime(debut_str, '%d/%m/%Y %H:%M')
            fin = datetime.strptime(fin_str, '%d/%m/%Y %H:%M')
            

            new_mission.debut = debut
            new_mission.fin = fin


            if isinstance(request.user, User):
                new_mission.envoyeur = request.user

            new_mission.encours = True


            # Calculer la durée de la mission en heures
            duree_mission = (new_mission.fin - new_mission.debut).total_seconds() / 3600

            if duree_mission >= 0:
                temps_actuel = timezone.now()
                if new_mission.debut <= temps_actuel <= new_mission.fin:
                    # La mission est en cours
                    new_mission.encours = True
                elif temps_actuel > new_mission.fin:
                    # La mission est terminée
                    new_mission.encours = False
                    new_mission.terminer = True
                else:
                    # La mission n'a pas encore commencé
                    new_mission.encours = False
            else:
                return HttpResponse('Date invalide')

            # Enregistrez la mission dans la base de données
            new_mission.save()

            # Redirigez l'utilisateur vers la page souhaitée après la création de la mission
            return redirect('mission')

    else:
        form = CreateMissionForm()

    return render(request, 'creer_mission.html', {'form': form})

def update_mission(request, my_id):
    obj = get_object_or_404(mission, id=my_id)
    form = CreateMissionForm(request.POST or None, instance=obj)
    if request.method == 'POST':
        if form.is_valid():
            mission_instance = form.save(commit=False)
            
            mission_instance.save()
            return redirect('voir_mission')
        else:
            print(form.errors)
    return render(request, 'update_mission.html',  {'form': form})

def voir_mission(request):
    obj = mission.objects.all()
    return render(request,"voir_mission.html", {'obj':obj})

@csrf_exempt
def delete_mission(request, my_id):
    missions = get_object_or_404(mission, id=my_id)
    if request.method == "DELETE":
        missions.delete()
        return HttpResponse(status=204)

    return render(request, 'delete_mission.html')

def create_message(request):
    if request.method == 'POST':
        form = CreateMessageForm(request.POST)
        if form.is_valid():
            # Créer une instance de mission à partir des données du formulaire
            new_message = form.save(commit=False)
            
            if isinstance(request.user, User):
                new_message.envoyeur = request.user

            # Enregistrez la mission dans la base de données
            new_message.save()

            return redirect('message')

    else:
        form = CreateMessageForm()

    return render(request, 'creer_message.html', {'form': form})

def voir_message(request):
    obj = message.objects.all()
    return render(request,"voir_message.html", {'obj':obj})

def upload_photo(request, my_id):
    obj = get_object_or_404(User, id=my_id)
    form = uploadImageForm(request.POST , request.FILES, instance=obj)
    if request.method == 'POST' or 'FILES':
        if form.is_valid():
            user_instance = form.save(commit=False)
            
            user_instance.save()
        else:
            print(form.errors)
    return render(request, 'upload_photo.html',  {'form': form})


#les vu des api_rest

#api d'inscription

class RegisterView(generics.CreateAPIView):
    serializer_class = RegisterSerializer

    def create(self, request, *args, **kwargs):
        print(request.data)
        # Utilisez le sérialiseur pour valider les données et créer un nouvel utilisateur
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
    
        user = serializer.save()
        # Vous pouvez ajouter d'autres logiques ici, par exemple, générer un token d'authentification.

        return Response({
            "user": serializer.data,
            "message": "L'utilisateur a été créé avec succès."
        }, status=201)

#api de connection
class LoginView(GenericAPIView):
    permission_classes = [AllowAny]
    serializer_class = LoginSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        username = serializer.validated_data['username']
        password = serializer.validated_data['password']
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)

            matricule = user.matricule
            travaille = user.heure_Travaille
            print(travaille)
            
            # Déterminez le statut de l'utilisateur en fonction de vos champs personnalisés
            if user.is_superuser : 
                user_status = 'admin'
            elif user.is_staff : 
                user_status = 'Controlleur de burreau'    
            elif user.is_terrain:
                user_status = 'Controlleur de Terrain'
            elif user.is_motoman:
                user_status = 'Motoman'
            else:
                user_status = 'Normal User'
            print(matricule)
            # Renvoyez la réponse JSON avec le statut
            return Response({'message': 'Logged in successfully.', 'statut': user_status, 'id': user.id, 
            'username':user.username, 'matricule':matricule, 'travaille': travaille})

        return Response({'message': 'Invalid credentials.'}, status=401)


    
#api d'envoie des rapport 

class RapportCreateView(generics.CreateAPIView):
    serializer_class = RapportSerializer

    def create(self, request, *args, **kwargs):
        # Affichez la requête JSON reçue depuis Flutter 
        print("Données reçues  :")
        print(request.data)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Affichez les données validées par le serializer
        print("Données validées par le serializer :")
        print(serializer.validated_data)

        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)

        return Response(serializer.data, headers=headers)
        
class UserListAPIView(generics.ListAPIView):
    queryset = User.objects.filter(is_superuser=True) | User.objects.filter(is_staff=True)
    serializer_class = UserSerializer

class UserListAllAPIView(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

class RapportListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, envoyeur_id):  # Récupérer l'envoyeur_id de l'URL
        rapports = Rapport.objects.filter(envoyeur__id=envoyeur_id)  # Filtrer les rapports par l'ID de l'envoyeur
        serializer = RapportListSerializer(rapports, many=True)
        return Response(serializer.data)



class PositionCreateView(generics.CreateAPIView):
    serializer_class = PositionSerailize

#mission

class missionListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, receveur_id):
        missions = mission.objects.filter(receveur__id=receveur_id)
        serializer = missionSerializer(missions, many=True)
        
        # Ajoutez mission_status et mission_status2 au contexte du serializer
        for mission_instance, data in zip(missions, serializer.data):
            data['mission_status'] = mission_instance.get_mission_status()
            data['mission_status2'] = mission_instance.get_mission_status2()

        return Response(serializer.data)
#pour compter les message, mission et point
class MissionPointListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, receveur_id):
        missions = mission.objects.filter(receveur__id=receveur_id, reussie=True)
        points_sum = missions.aggregate(total_points=Sum('point'))['total_points']
        messages = message.objects.filter(receveur__id=receveur_id).count()
        mission_count = mission.objects.filter(receveur__id=receveur_id).count()
        
        return Response({'total_points': points_sum, 'message_count': messages, 'mission_count':mission_count,
        'mission_sucess_count':missions})


class messageListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, receveur_id):
        messages = message.objects.filter(receveur__id=receveur_id)
        serializer = messageSerializer(messages, many=True)


        return Response(serializer.data)

class UserImageView(generics.RetrieveUpdateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer_image
    lookup_field = 'id'